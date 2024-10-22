import fitz
from openpyxl import load_workbook


def get_rgb_color(color_name):
    color_map = {
        "light blue": (173, 216, 230),  
        # Add more color mappings as needed
    }
    return color_map.get(color_name.lower(), (255, 255, 255))  # Default to black if color not found



def is_non_annotatable_region(page, x1, y1, x2, y2, search_text):
    # We don't want to annotate text within 100 units from the top of the page
    if y1 < 100:
        return True
    
    # Check if there are words before the search text
    # We want to ensure the search text is on a line by itself
    # We can check if there are any non-whitespace characters before the search text within a certain distance
    distance_threshold = 20  # Adjust this threshold as needed
    text_before_search_text = page.get_text("text", clip=(x1 - distance_threshold, y1, x1, y2))
    if any(word.strip() for word in text_before_search_text.split() if word.strip()):
        return True
    
    return False






def find_text_coordinates(pdf_path, search_text, excel_file):
    # Open the PDF file
    doc = fitz.open(pdf_path)

    text_coordinates = []

    # Iterate through each page of the PDF
    for page_num in range(len(doc)):
        page = doc[page_num]

        # Search for the given text on the current page
        for match in page.search_for(search_text):
            # Append the coordinates of the text match to the list
            text_coordinates.append((match[0], match[1], match[2], match[3], page_num + 1))
    
    doc.close()

    return text_coordinates





def draw(page, search_text, box_properties):
    # Search for instances of the given text on the current page
    text_instances = page.search_for(search_text)
    print(f"Found {len(text_instances)} instances of '{search_text}' ")

    # Iterate through each instance of the text found on the page
    for match in text_instances:
        x1, y1, x2, y2 = match

        print(f"Text instance coordinates: ({x1}, {y1}), ({x2}, {y2})")
        box_text = box_properties['text']
        
        # Skip annotation if text is found within certain regions (e.g., titles)
        if is_non_annotatable_region(page, x1, y1, x2, y2,box_text ):
            print("Text instance found in non-annotatable region. Skipping...")
            continue    
        
        # Calculate the center of the text box
        center_x = ((x1 + x2)+150) / 2
        center_y = (y1 + y2) / 2

        text_width = len(box_properties['text']) * 10

        left_x = center_x - text_width / 2
        right_x = center_x + text_width / 2
        # Calculate the width and height of the text box
        width = x2 - x1
        height = y2 - y1
        # Adjust the rectangle coordinates to create the annotation box
        rect = fitz.Rect(int(left_x), int(y1), int(right_x), int(y2))
        print(f"Annotation box coordinates: {(rect)}")
        color = box_properties['color']


        try:
            # Convert color to RGB integers
            # color_int = get_rgb_color(color)
            # Add annotation with converted RGB color values
            annot = page.add_freetext_annot(rect, box_text) 
            annot.update(fontsize=10, fontname='helv', fill_color=color)
            print("Annotation added successfully.")
        except Exception as e:
            print(f"An error occurred while adding annotation: {e}")
 



if __name__ == "__main__":
    try:
        wb = load_workbook(filename='Excel sheet')
        sheet = wb.active

        max_row = sheet.max_row

        # Input parameters
        pdf_path = "Path to PDF you want to Annotate"
        output_path = "Path to new pdf created with Annotations"

        doc = fitz.open(pdf_path)

        search_texts = []
        box_properties_list = []  # List to store box properties for each search text

        for row in range(2, max_row + 1):
            search_text = str(sheet.cell(row=row, column=2).value)
            box_properties = {
                'text': str(sheet.cell(row=row, column=4).value),
                'color': (173/255, 216/255, 230/255)
            }

            # Check if search_text is a string
            if isinstance(search_text, str):
                # If it's a string, strip leading and trailing whitespace, parentheses, and commas
                search_text = search_text.strip('(),')
            else:
                print("Value from Excel sheet is not a string:", search_text)

            # Append search_text to the list
            search_texts.append(search_text)
            # Append box_properties to the list
            box_properties_list.append(box_properties)


        for page in doc:
            print(f"Proccessing page {page.number}")
            for search_text, box_properties in zip(search_texts, box_properties_list):
                print("Search Text:", search_text)
                draw(page, search_text, box_properties)
        doc.save(output_path)
        print("Document saved successfully.")
        doc.close()

    except Exception as e:
        print("An error occurred:", e)